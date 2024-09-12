from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.conf import settings
import os
import openpyxl
from .models import Company, Factory, Product, FactoryProduct
from .serializers import FileUploadSerializer

def process_excel_and_create_entries(file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Usar a primeira planilha

    # Obter os nomes das fábricas a partir das células específicas
    factory_names = {
        'H1': sheet['H1'].value,  # Fábrica H1
        'L1': sheet['L1'].value,  # Fábrica L1
        'P1': sheet['P1'].value,  # Fábrica P1
        'T1': sheet['T1'].value,  # Fábrica T1
        'AA1': sheet['AA1'].value  # Fábrica AA1
    }

    # Criar ou obter as fábricas
    factories = {}
    for cell, factory_name in factory_names.items():
        if factory_name:
            factory, _ = Factory.objects.get_or_create(name=factory_name)
            factories[cell] = factory

    # Obter os nomes das companhias a partir das células específicas (linhas de empresas)
    company_columns = {
        'H3': {'name': sheet['H3'].value, 'sim_col': 'H'},  # Coluna para SIM verificação
        'I3': {'name': sheet['I3'].value, 'sim_col': 'I'},
        'L3': {'name': sheet['L3'].value, 'sim_col': 'L'},
        'M3': {'name': sheet['M3'].value, 'sim_col': 'M'},
        'P3': {'name': sheet['P3'].value, 'sim_col': 'P'},
        'T3': {'name': sheet['T3'].value, 'sim_col': 'T'},  # Empresas separadas por '/'
        'U3': {'name': sheet['U3'].value, 'sim_col': 'U'},
        'V3': {'name': sheet['V3'].value, 'sim_col': 'V'},
        'X3': {'name': sheet['X3'].value, 'sim_col': 'X'},
        'AA3': {'name': sheet['AA3'].value, 'sim_col': 'AA'},
        'AB3': {'name': sheet['AB3'].value, 'sim_col': 'AB'}
    }

    # Criar ou obter as empresas (Company)
    companies = {}
    for cell, company_info in company_columns.items():
        company_name = company_info['name']
        if company_name:
            # Algumas colunas têm múltiplas empresas separadas por "/"
            company_names = [name.strip() for name in company_name.split('/')]
            for name in company_names:
                company, _ = Company.objects.get_or_create(name=name)
                companies[name] = company
                print(f"Empresa processada: {name}")

    # Definir as colunas das fábricas e pesos corretamente
    factory_columns = {
        'H1': {'factory_code_col': 'J', 'weight_col': 'K', 'start_row': 4},  # Fábrica H1
        'L1': {'factory_code_col': 'N', 'weight_col': 'O', 'start_row': 4},  # Fábrica L1
        'P1': {'factory_code_col': 'Q', 'weight_col': 'R', 'start_row': 4},  # Fábrica P1
        'T1': {'factory_code_col': 'Y', 'weight_col': 'Z', 'start_row': 6},  # Fábrica T1
        'AA1': {'factory_code_col': 'AC', 'weight_col': 'AD', 'start_row': 4}  # Fábrica AA1
    }

    # Processar as linhas da planilha a partir da linha 4 para criar produtos e associá-los
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        ncm = row[1].value  # Coluna B (NCM)
        alumifont_code = row[3].value  # Coluna D (alumifont_code)
        image = row[4].value  # Coluna E (imagem)
        length_mm = row[5].value  # Coluna F (length_mm)

        # Verificar se o alumifont_code é válido e não nulo
        if not alumifont_code:
            print(f"Erro: Alumifont Code está ausente na linha {row[0].row}.")
            continue

        # Criar ou obter o produto
        product, created = Product.objects.get_or_create(
            alumifont_code=alumifont_code,
            defaults={
                'ncm': ncm,
                'image': image,
                'length_mm': length_mm if length_mm else 6000,
                'temper_alloy': '6063T5',
                'surface_finish': None
            }
        )

        # Relacionar o produto com as fábricas
        for factory_cell, columns in factory_columns.items():
            current_row = row[0].row
            if current_row < columns['start_row']:
                continue  # Ignorar se a linha atual for menor que a linha inicial da fábrica

            factory_code_col = columns['factory_code_col'] + str(current_row)
            weight_col = columns['weight_col'] + str(current_row)

            factory_code = sheet[factory_code_col].value
            weight_m_kg = sheet[weight_col].value

            # Verificar se o código da fábrica e o peso estão preenchidos
            if factory_code and weight_m_kg:
                try:
                    factory = factories.get(factory_cell)
                    if factory:
                        # Verifique se `factory_code` é uma string antes de usar `strip()`
                        if isinstance(factory_code, str):
                            factory_code = factory_code.strip()

                        # Substituir vírgula por ponto para lidar com números decimais
                        weight_m_kg = str(weight_m_kg).replace(',', '.')

                        # Criar ou atualizar a entrada FactoryProduct
                        FactoryProduct.objects.update_or_create(
                            factory=factory,
                            product=product,
                            defaults={
                                'factory_code': factory_code,
                                'weight_m_kg': float(weight_m_kg)
                            }
                        )
                        print(f"Produto {alumifont_code} associado à fábrica {factory_cell} com código {factory_code} e peso {weight_m_kg}.")
                except Exception as e:
                    print(f"Erro ao processar a fábrica {factory_cell} para o produto {alumifont_code}: {e}")
            else:
                print(f"Produto {alumifont_code}: Fábrica {factory_cell} não oferece este produto.")

        # Verificar empresas habilitadas ("SIM") para este produto e criar relação
        for company_col, company_info in company_columns.items():
            enabled_value = sheet[company_info['sim_col'] + str(row[0].row)].value
            if enabled_value and enabled_value.strip().upper() == "SIM":
                company_name = company_info['name']
                company = companies.get(company_name)
                if company:
                    # Certifique-se de que a empresa é adicionada ao Many-to-Many `enabled_companies`
                    product.enabled_companies.add(company)
                    print(f"Produto {alumifont_code} habilitado para a companhia: {company_name}")

                    # Relacionar a empresa à(s) fábrica(s) correspondente(s)
                    for factory_cell, factory in factories.items():
                        if factory_code and factory:
                            company.factories.add(factory)  # Adiciona a fábrica à empresa
                            print(f"Empresa {company_name} associada à fábrica {factory.name}")

        # Salvar o produto após a atualização
        product.save()

class ExcelUploadViewSet(viewsets.ViewSet):
    serializer_class = FileUploadSerializer
    parser_classes = (MultiPartParser, FormParser)

    def create(self, request, *args, **kwargs):
        """
        POST /order-planilhas/ : Upload e processamento de planilha Excel
        """
        file_obj = request.FILES.get('file', None)
        if not file_obj:
            return Response({"detail": "Nenhum arquivo foi enviado."}, status=status.HTTP_400_BAD_REQUEST)

        # Definir o caminho completo para o arquivo de upload
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
        os.makedirs(upload_dir, exist_ok=True)  # Criar o diretório, se não existir
        
        file_path = os.path.join(upload_dir, file_obj.name)

        # Salvar o arquivo no diretório de upload
        with open(file_path, 'wb+') as destination:
            for chunk in file_obj.chunks():
                destination.write(chunk)

        # Verificar se o arquivo foi salvo corretamente
        if not os.path.exists(file_path):
            return Response({"detail": f"O arquivo não foi encontrado após o upload: {file_path}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Processar o arquivo Excel
        try:
            process_excel_and_create_entries(file_path)
        except Exception as e:
            return Response({"detail": f"Erro ao processar o arquivo: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"detail": "Arquivo processado com sucesso."}, status=status.HTTP_201_CREATED)



            


class ExcelUploadViewSet(viewsets.ViewSet):
    serializer_class = FileUploadSerializer
    parser_classes = (MultiPartParser, FormParser)

    def create(self, request, *args, **kwargs):
        """
        POST /order-planilhas/ : Upload e processamento de planilha Excel
        """
        file_obj = request.FILES.get('file', None)
        if not file_obj:
            return Response({"detail": "Nenhum arquivo foi enviado."}, status=status.HTTP_400_BAD_REQUEST)

        # Definir o caminho completo para o arquivo de upload
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
        os.makedirs(upload_dir, exist_ok=True)  # Criar o diretório, se não existir
        
        file_path = os.path.join(upload_dir, file_obj.name)

        # Salvar o arquivo no diretório de upload
        with open(file_path, 'wb+') as destination:
            for chunk in file_obj.chunks():
                destination.write(chunk)

        # Verificar se o arquivo foi salvo corretamente
        if not os.path.exists(file_path):
            return Response({"detail": f"O arquivo não foi encontrado após o upload: {file_path}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Processar o arquivo Excel
        try:
            process_excel_and_create_entries(file_path)
        except Exception as e:
            return Response({"detail": f"Erro ao processar o arquivo: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"detail": "Arquivo processado com sucesso."}, status=status.HTTP_201_CREATED)
