import os
import openpyxl
from django.core.files import File
from django.core.management.base import BaseCommand
from PedidoAPP.models import Product
from django.conf import settings

class Command(BaseCommand):
    help = 'Import products from an XLSX file'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_file', type=str)

    def handle(self, *args, **kwargs):
        xlsx_file = kwargs['xlsx_file']
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active

        # Define the starting row and the ending row
        start_row = 20
        end_row = 236

        for row in range(start_row, end_row + 1):
            alumifont_code = sheet.cell(row=row, column=3).value  # Column C
            factory_code = sheet.cell(row=row, column=4).value  # Column D
            image_path = sheet.cell(row=row, column=5).value  # Column E (assuming this is a file path)
            length_mm = sheet.cell(row=row, column=6).value  # Column F
            temper_alloy = sheet.cell(row=row, column=7).value  # Column G
            weight_m_kg = sheet.cell(row=row, column=8).value  # Column H

            # Handle empty or invalid values if needed
            if not all([alumifont_code, factory_code, length_mm, temper_alloy, weight_m_kg]):
                continue

            image_file = None
            if image_path:
                # Construct the full path to the image file
                full_image_path = os.path.join(settings.MEDIA_ROOT, 'product_images', os.path.basename(image_path))
                if os.path.isfile(full_image_path):
                    with open(full_image_path, 'rb') as img_file:
                        image_file = File(img_file, name=os.path.basename(full_image_path))

            # Create the Product object
            Product.objects.create(
                alumifont_code=alumifont_code,
                factory_code=factory_code,
                image=image_file,  # This should be the File object or None
                length_mm=length_mm,
                temper_alloy=temper_alloy,
                weight_m_kg=weight_m_kg
            )

        self.stdout.write(self.style.SUCCESS('Successfully imported products'))
